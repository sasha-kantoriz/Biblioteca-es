import re
import fitz
import fpdf
import openpyxl
import requests
import os
import argparse
from time import sleep
from openai import OpenAI
from os.path import expanduser
from pathlib import Path
from datetime import datetime
from urllib.parse import urlparse
from urllib.parse import parse_qs
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.chrome.options import Options


global client
global driver
global book_id
global start_index 
global end_index
global outdir

book_id = 1
outdir = datetime.now().strftime('%Y-%B-%d')
Path(outdir).mkdir(parents=True, exist_ok=True)
Path(f'{outdir}/interior').mkdir(parents=True, exist_ok=True)
Path(f'{outdir}/cover').mkdir(parents=True, exist_ok=True)


# ==========================================================================
# PDF
class PDF(fpdf.FPDF):
    def footer(self):
        if self.page_no() != 1:
            # Go to 1.5 cm from bottom
            self.set_y(-15)
            self.add_font("dejavu-sans", style="", fname="assets/DejaVuSans.ttf")
            self.set_font("dejavu-sans", size=12)
            # Print centered page number
            self.cell(0, 10, f"{self.page_no()}", 0, 0, 'C')


# ==========================================================================
# EXCEL
# ==========================================================================
try:
    wb = openpyxl.load_workbook('Project-Biblioteca.xlsx')
except:
    wb = openpyxl.Workbook()
try:
    del wb['Sheet']
except:
    pass
datestamp = datetime.now().strftime('%Y-%B-%d %H_%M')
ws = wb.create_sheet(datestamp)
ws.append(["Book ID", "Origin URL", "Title", "Language",
           "Author", "Translator", "Illustrator",
           "Description", "Keywords", "BISAC codes",
           "Pages num", "PDF file name", "Cover PDF file name"])


# ==========================================================================
# text formatting
def format_book_text(text):
    formatted = text.replace('©', '').replace('Biblioteca Nacional de España', '')
    formatted = formatted.replace('«', '"').replace('»', '"')
    formatted = formatted.replace('\n\n\n\n', '\n\n').replace('\n\n\n', '\n\n')
    formatted = formatted.replace('\n\n', '____').replace('\n', '').replace('____', '\n\n')
    formatted = formatted.replace('*', '')\
        .replace('/', '')\
        .replace("\\", '')\
        .replace('|', '')\
        .replace('¿', '')\
        .replace('•', '')\
        .replace('<', '')\
        .replace('>', '')\
        .replace('~', '')\
        .replace('^', '')\
        .replace('Г', '')\
        .replace('%', '')\
        .replace('=', '')\
        .replace('$', '')\
        .replace('ϕ', '')\
        .replace(' ,', ',')\
        .replace(' ;', ';')\
        .replace(' :', ':')\
        .replace(' !', '!')\
        .replace(' ?', '?')\
        .replace(' .', '.')\
        .replace(' -', '-')\
        .replace('-;', ';')\
        .replace('";', ';')\
        .replace("';", '')\
        .replace('",', ',')\
        .replace('-,', ',')\
        .replace('--,', '-')\
        .replace('-.', '.')\
        .replace(';-', ';')\
        .replace('-.', '.')\
        .replace('.,', ',')\
        .replace('.:', ':')\
        .replace(' ( )', '')
    # .replace("-'", "'") \
    # .replace('."', '.')\
    # - "
    # regexp for removing dashes inside words
    formatted = re.sub(r'(\w)-(\w)', r'\1\2', formatted)
    # regexp for removing dots from beggining of sentences
    formatted = re.sub(r'\n\.', r'\n', formatted)
    # regexp for removing dashes from beggining of sentences
    formatted = re.sub(r'\n-', r'\n', formatted)
    # regexp for removing commas from beggining of sentences
    formatted = re.sub(r'\n,', r'\n', formatted)
    # regexp for removing random? numbers between sentences
    formatted = re.sub(r'\n\d+\s?\n', r'', formatted)
    return formatted
# ==========================================================================


def generate_book_pdfs(text, url, title, author, language='es'):
    global outdir
    interior_pdf_fname, cover_pdf_fname = f"{book_id}_paperback_interior.pdf", f"{book_id}_paperback_cover.pdf"
    # Interior
    pdf = PDF(format=(152.4, 228.6))
    pdf.add_font("dejavu-sans", style="", fname="assets/DejaVuSans.ttf")
    # TITLE
    if title and author:
        pdf.add_page()
        pdf.set_font("dejavu-sans", size=12)
        # center vertically first page text
        lines_num = len(pdf.multi_cell(w=0, align='C', padding=(0, 8), text=f"{title}\n\n{author}", dry_run=True, output="LINES"))
        if lines_num >= 3:
            padding_top = (228.6 - 24 * (lines_num - 1)) / 2
        else:
            padding_top = (228.6 - 24 * (lines_num)) / 2
        pdf.multi_cell(w=0, align='C', padding=(padding_top, 8, 0), text=f"{title}\n\n{author}")
    # TEXT
    pdf.add_page()
    pdf.set_font("dejavu-sans", size=12)
    pdf.multi_cell(w=0, h=4.6, align='J', padding=8, text=text)
    # FOOTER
    pdf.add_font('dejavu-sans', style="", fname="assets/DejaVuSans.ttf")
    pdf.add_page()
    pdf.multi_cell(w=0, h=4.6, align='J', padding=8, text="""Este libro incluye imágenes y/o contenidos obtenidos de los fondos de la Biblioteca Nacional de España, disponibles en la Biblioteca Digital Hispánica y la Hemeroteca Digital a través del sitio web bne.es. Estas imágenes y contenidos están en dominio público y se utilizan bajo una licencia de Reconocimiento 4.0 Internacional de Creative Commons. El uso de estas imágenes y contenidos es gratuito y no ha requerido autorización previa, siendo aplicable tanto para fines no comerciales como comerciales y académicos. Al utilizar estas imágenes y contenidos, reconocemos y citamos la procedencia de los mismos como: “Imágenes procedentes de los fondos de la Biblioteca Nacional de España".""")
    # check book size
    pages = pdf.page_no()
    if 24 <= pages <= 828:
        # Render book interior
        pdf.output(f"{outdir}/interior/{interior_pdf_fname}")
        # BOOK COVER
        cover_width, cover_height = 152.4 * 2 + pages * 0.05720 + 3.175 * 2, 234.95
        pdf = fpdf.FPDF(format=(cover_width, cover_height))
        pdf.add_font('dejavu-sans', style="", fname="assets/DejaVuSans.ttf")
        pdf.add_page()
        pdf.set_fill_color(r=250,g=249,b=222)
        pdf.rect(h=pdf.h, w=pdf.w, x=0, y=0, style="DF")
        cols = pdf.text_columns(ncols=2, gutter=pages*0.05720 + 1.588*2, l_margin=6.35, r_margin=6.35)
        #
        description_query = f"Provide a 150 words description of the classic book {title}"
        if author:
            description_query += f" by Author and Writer {author}."
        if language:
            description_query += f" Write the review in this language: {language}"
        description_completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": description_query
                },
            ]
        )
        description = description_completion.choices[0].message.content
        #
        keywords_query = f'Give me 7 keywords separated by semicolons (only the keywords, no numbers nor introductory workds) for the classic book "{title}" by Author "{author}". Keywords must not be subjective claims about its quality, time-sensitive statments and must not include the word "book". Keywords must also not contain words included on the book the title, author nor contained on the following book description: {description}'
        keywords_completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": keywords_query
                },
            ]
        )
        keywords = keywords_completion.choices[0].message.content
        #
        bisac_codes_query = f'Give me up to 3 BISAC codes separated by semicolons (only the code, not its description and not numbered) for the book "{title}" by Author "{author}" with description "{description}", for its correct classification. Output format example would be: FIC019000; FIC031010; FIC014000'
        bisac_codes_completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {
                    "role": "system",
                    "content": bisac_codes_query
                },
            ]
        )
        bisac_codes = bisac_codes_completion.choices[0].message.content
        #
        description_p = cols.paragraph(text_align='L')
        pdf.set_font('dejavu-sans', size=12)
        description_lines = pdf.multi_cell(w=152.4, align='L', padding=(0, 11.175), text=description, dry_run=True, output="LINES")
        description_p.write('\n'.join(description_lines))
        cols.end_paragraph()
        #
        cols.new_column()
        #
        title_p = cols.paragraph(text_align='C')
        pdf.set_font('dejavu-sans', size=28)
        title_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=f"\n\n{title}", dry_run=True, output="HEIGHT")
        title_p.write(f"\n\n{title}")
        cols.end_paragraph()
        #
        separator_text = "\n* * *"
        separator_p = cols.paragraph(text_align='C')
        pdf.set_font('dejavu-sans', size=16)
        separator_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=separator_text, dry_run=True, output="HEIGHT")
        separator_p.write(separator_text)
        cols.end_paragraph()
        #
        author_p = cols.paragraph(text_align='C')
        pdf.set_font('dejavu-sans', size=16)
        author_h = pdf.multi_cell(w=0, align='C', padding=(0, 8), text=f"\n{author}", dry_run=True, output="HEIGHT")
        author_p.write(f"\n{author}")
        cols.end_paragraph()
        #
        cols.render()
        #
        pdf.output(f"{outdir}/cover/{cover_pdf_fname}")
        #
        ws.append([book_id, url, title, 'es', author, '', '',
                   description, keywords, bisac_codes, pages, f"{outdir}/interior/{interior_pdf_fname}", f"{outdir}/cover/{cover_pdf_fname}"])


def download_books_per_page(driver: webdriver):
    global book_id, start_index, end_index

    books = driver.find_elements(By.CSS_SELECTOR, "#lista div div.details h2 a")
    # Main loop
    for i in range(len(books)):
        book = books[i]
        #
        if end_index and book_id > end_index:
            return "completed"
        #
        if start_index <= book_id:
            book_download_url = "https://bdh-rd.bne.es/high.raw?id={_id}&name=00000001.original.pdf&view=main&lang=es"
            book_url = book.get_attribute("href")
            driver.execute_script("window.open('%s', '_blank')" % book_url)
            #
            details_tab = None
            for tab in driver.window_handles:
                if tab != books_list_tab:
                    details_tab = tab
                    break
            driver.switch_to.window(details_tab)
            #
            try:
                title = driver.find_element(By.XPATH, '//*[@id="results"]/div[1]/div/div[2]/h1').text
                title = title.replace('[Texto impreso]', '').strip()
            except:
                title = ''
            try:
                author = driver.find_element(By.XPATH, '//*[@id="results"]/div[1]/div/div[2]/h2').text
                author = author.split(',')[0]
            except:
                author = ''
            #
            download_link_element = driver.find_element(By.XPATH, '//*[@id="results"]/div[1]/div/div[1]/div[2]/a')
            download_link = download_link_element.get_attribute('href')
            parsed_url = urlparse(download_link)
            captured_id = parse_qs(parsed_url.query)['id'][0]
            r = requests.get(book_download_url.format(_id=captured_id), verify=False)
            with open(f'tmp/{book_id}.pdf', 'wb') as f:
                f.write(r.content)
            #
            driver.close()
            driver.switch_to.window(books_list_tab)
            #
            pdf_file, book_text = fitz.open(f'tmp/{book_id}.pdf'), ""
            for page in pdf_file:
                blocks = page.get_text('blocks', flags=fitz.TEXT_INHIBIT_SPACES | fitz.TEXT_PRESERVE_LIGATURES | fitz.TEXT_PRESERVE_WHITESPACE | fitz.TEXT_PRESERVE_SPANS | fitz.TEXT_MEDIABOX_CLIP)
                for b in blocks:
                    book_text += "\n" + b[4]
            # format book text
            book_text = format_book_text(book_text)
            # generate PDFs
            generate_book_pdfs(book_text, book_url, title, author)

        book_id += 1
        books = driver.find_elements(By.CSS_SELECTOR, "#lista div div.details h2 a")

    return "success"


if __name__ == '__main__':
    Path('tmp').mkdir(parents=True, exist_ok=True)
    #
    parser = argparse.ArgumentParser()
    parser.add_argument("-s","--start_index", help="", type=int, default=0)
    parser.add_argument("-e","--end_index", help="", type=int, default=None)
    args = parser.parse_args()
    start_index = args.start_index
    end_index = args.end_index
    #
    client = OpenAI()
    #
    chrome_options = Options()
    chrome_options.enable_downloads = True
    chrome_options.add_argument("--window-size=1920x1080")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--verbose')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": "./downloads/",
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing_for_trusted_sources_enabled": False,
        "safebrowsing.enabled": False
    })

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

    retries = 10
    while retries:
        try:
            driver.get('https://bdh.bne.es/bnesearch/AdvancedSearch.do?showAdvanced=true')
            element = driver.find_element(By.XPATH, '//*[@id="btn_busqueda_avanzada"]')
            if element.is_displayed():
                break
        except (NoSuchElementException, StaleElementReferenceException):
            if retries <= 0:
                raise
            else:
                driver.refresh()
        retries = retries - 1
        sleep(4)

    driver.find_element(By.XPATH, '//*[@id="field1"]').click()
    driver.find_element(By.XPATH, '//*[@id="field1"]/option[12]').click()
    driver.find_element(By.XPATH, '//*[@id="AdvancedSearch"]/fieldset[6]/div[2]/ul/li[2]/input').click()
    driver.find_element(By.XPATH, '//*[@id="btn_busqueda_avanzada"]').click()

    books_list_tab = driver.current_window_handle

    # process first results page
    if download_books_per_page(driver) == "completed":
        wb.save("Project-Biblioteca.xlsx")
        exit(0)
    # process second results page
    driver.find_element(By.XPATH, '//*[@id="navsup"]/span[5]/a/img').click()
    if download_books_per_page(driver) == "completed":
        wb.save("Project-Biblioteca.xlsx")
        exit(0)
    driver.find_element(By.XPATH, '//*[@id="navsup"]/span[10]/a/img').click()
    # process in a loop by pages
    navsup_element = driver.find_element(By.XPATH, '//*[@id="navsup"]/span[11]/a/img')
    try:
        while navsup_element.is_displayed():
            if download_books_per_page(driver) == "completed":
                wb.save("Project-Biblioteca.xlsx")
                exit(0)
            navsup_element.click()
            navsup_element = driver.find_element(By.XPATH, '//*[@id="navsup"]/span[11]/a/img')
    except Exception as e:
        print(str(e))
